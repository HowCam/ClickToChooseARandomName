# TO-DO: Loading window for auto updating.
# DONE: 
#        改进了初次打开软件时的提示窗口。
#           -from:Can not choose from an empty sequence.
#           - to: 当前文件路径为空，请导入名单！
import os
import tkinter as tk
import random
import openpyxl
import keyboard
import requests
import webbrowser
import pygame
import subprocess
from tkinter import messagebox, filedialog

class FloatingWindow:
    def __init__(self, master):

        # 设置窗口
        self.master = master
        master.title("点名器v2.0.1")   # 设置窗口标题
        master.geometry("650x250+100+400")  # 设置窗口大小和位置
        bg_color = "#ccffcc"  # 存储默认背景颜色
        master.configure(bg=bg_color)  # 设置窗口背景颜色

        # 设置图标
        master.iconbitmap("icon.ico")

        # 存储当前版本号
        self.current_version = "2.0.1"

        # 获取github上的最新版本号和release描述
        self.latest_version, self.latest_version_desc = self.get_latest_version()

        # 创建姓名标签（居中）
        self.name_label = tk.Label(master, font=("楷体", 120), fg="#005f35", bg=bg_color)
        self.name_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # 创建显示文件路径的标签（左上）
        self.file_path_label = tk.Label(master, text="", font=("等线", 10), fg='gray', bg=bg_color)
        self.file_path_label.place(relx=0, rely=0, anchor=tk.NW)

        # 创建导入名单按钮（右上）
        self.import_button = tk.Button(master, text="导入名单", command=self.manual_import)
        self.import_button.pack(side=tk.TOP, anchor=tk.NE)

        # 添加提示文本（右下）
        self.tip_label = tk.Label(master, text="按下 V 键隐藏/显示窗口", font=("仿宋", 10), fg="gray", bg=bg_color)
        self.tip_label.pack(side=tk.BOTTOM, anchor=tk.SE)

        # 创建查看帮助按钮（右下）
        self.help_button = tk.Button(master, text="查看帮助", command=self.show_help)
        self.help_button.pack(side=tk.RIGHT, anchor=tk.SE)
        
        # 如果当前版本不是最新版本，则创建查看更新按钮（右下）
        if self.current_version != self.latest_version:
            self.update_button = tk.Button(master, text="查看更新", command=self.check_update)
            self.update_button.pack(side=tk.RIGHT, anchor=tk.SE)
        
        # 创建显示姓名总数的标签（左下）
        self.total_names_label = tk.Label(master, text="姓名总数：0", font=("仿宋", 10), bg=bg_color)
        self.total_names_label.place(relx=0, rely=1, anchor=tk.SW)

        # 创建音效开关按钮（左下）
        self.sound_button = tk.Button(master, text="音效：开", command=self.toggle_sound)
        self.sound_button.place(relx=0, rely=0.93, anchor=tk.SW)
        self.sound_enabled = True  # 初始音效状态为开启

        # 显示当前版本和最新版本及其描述（下居中）
        if self.current_version == self.latest_version:
            self.version_label = tk.Label(master, text=f"当前已是最新版本：{self.current_version}", font=("仿宋", 10), bg=bg_color)
        else:
            self.version_label = tk.Label(master, text=f"当前版本：{self.current_version} 最新版本：{self.latest_version}", font=("仿宋", 10), bg=bg_color)
        self.version_label.place(relx=0.5, rely=1.0, anchor=tk.S)

        # 加载姓名数据
        self.load_names_from_excel()

        # 绑定键盘/鼠标触发刷新
        self.name_label.bind("<Button-1>", self.update_name)
        self.master.bind("<space>", self.update_name)
        self.master.bind("<Return>", self.update_name)
        self.master.bind("<Down>", self.update_name)

        # 添加全局键盘监听事件
        self.visible_pressed = False
        keyboard.on_press_key("v", self.toggle_window)

        # 初始化pygame
        pygame.mixer.init()

        # 设置窗口置顶
        master.attributes("-topmost", True)

    # 加载姓名数据
    def load_names_from_excel(self):
        try:
            # 尝试读取上次导入的文件路径
            with open("last_file_path.txt", "r") as f:
                file_path = f.read().strip()
        except FileNotFoundError:
            file_path = None

        if file_path:
            try:
                self.names = []
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    self.names.extend(row)
                self.file_path_label.config(text=f"当前文件路径：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"加载姓名数据失败：{e}")
        else:
            self.names = []
            self.file_path_label.config(text="当前文件路径：无")

        # 更新姓名总数标签
        self.total_names_label.config(text=f"姓名总数：{len(self.names)}")

    # 选择随机姓名
    def update_name(self, event=None):
        try:
            name = random.choice(self.names)
            self.name_label.config(text=name)
            # 播放音效
            if self.sound_enabled:
                pygame.mixer.music.load("sound.wav")
                pygame.mixer.music.play()
        except Exception as e:
            try:
            # 尝试读取上次导入的文件路径
                with open("last_file_path.txt", "r") as f:
                    file_path = f.read().strip()
            except FileNotFoundError:
                file_path = None
            
            if file_path == None:
                messagebox.showerror("错误", f"当前文件路径为空，请导入名单！")
            else:
                messagebox.showerror("错误", f"出现错误：{e}")

    # 切换隐藏/显示窗口状态
    def toggle_window(self, event):
        self.visible_pressed = not self.visible_pressed
        if self.visible_pressed:
            self.master.withdraw()  # 隐藏窗口
        else:
            self.master.deiconify()  # 显示窗口

    # 获取远程服务器上的最新版本号和release描述
    def get_latest_version(self):
        try:
            response = requests.get("https://api.github.com/repos/HowCam/ClickToChooseARandomName/releases/latest", timeout=5)
            latest_version = response.json()["tag_name"]
            latest_version_desc = response.json()["body"]
            return latest_version, latest_version_desc
        except Exception as e:
            messagebox.showerror("错误", f"获取最新版本号失败：{e}")
            return None, None

    # 手动导入功能
    def manual_import(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                self.names = []
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    self.names.extend(row)
                self.save_last_file_path(file_path)
                self.update_name()  # 导入成功后刷新一次姓名，避免导入后不出现姓名。
                self.file_path_label.config(text=f"当前文件路径：{file_path}")
                self.total_names_label.config(text=f"姓名总数：{len(self.names)}")  # 更新姓名总数
                messagebox.showinfo("提示", "导入成功！")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{e}")

    # 保存上次导入的文件路径
    def save_last_file_path(self, file_path):
        try:
            with open("last_file_path.txt", "w") as f:
                f.write(file_path)
        except Exception as e:
            messagebox.showerror("错误", f"保存文件路径失败：{e}")

    # 检查更新并显示通知
    def check_and_notify_update(self):
        if self.latest_version and self.latest_version != self.current_version:
            message = f"发现新版本 {self.latest_version}，是否下载并安装？\n（下载过程用时大约为1分钟，请勿关闭程序！）\n（下载完成后将自动关闭程序并运行安装包）\n\n版本描述：\n\n{self.latest_version_desc}"
            if messagebox.askyesno("更新提示", message):
                self.download_and_install_update()

    # 下载并安装更新
    def download_and_install_update(self):
        try:
            url = f"https://gitee.com/zhaofenghao/ClickToChooseARandomName/releases/download/{self.latest_version}/setup.exe"
            response = requests.get(url)
            with open("setup.exe", "wb") as f:
                f.write(response.content)
            # 执行安装文件
            subprocess.Popen(["setup.exe"])
            self.master.destroy()  # 关闭当前程序
        except Exception as e:
            messagebox.showerror("错误", f"下载和安装更新失败：{e}")

    # 查看更新，打开最新版本的 release 页面
    def check_update(self):
        webbrowser.open("https://github.com/HowCam/ClickToChooseARandomName/releases/latest")

    # 切换音效状态
    def toggle_sound(self):
        self.sound_enabled = not self.sound_enabled
        if self.sound_enabled:
            self.sound_button.config(text="音效：开")
        else:
            self.sound_button.config(text="音效：关")
    
    # 显示帮助信息，打开最新README.md文件
    def show_help(self):
        webbrowser.open("https://github.com/HowCam/ClickToChooseARandomName/blob/main/README.md")

def main():
    root = tk.Tk()
    app = FloatingWindow(root)
    app.update_name()  # 初始化时更新一次姓名，避免打开软件不显示姓名。
    app.check_and_notify_update()  # 检查更新
    root.mainloop()

if __name__ == "__main__":
    main()