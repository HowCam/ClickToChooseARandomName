import tkinter as tk
import random
import webbrowser
import openpyxl
import keyboard
import pygame
from tkinter import messagebox, filedialog

class FloatingWindow:
    def __init__(self, master):
        self.master = master
        master.title("点名器离线稳定版")   # 设置窗口标题
        master.geometry("650x240+100+400")  # 设置窗口大小和位置
        bg_color = "#ccffcc"  # 设置背景颜色
        master.configure(bg=bg_color)  # 设置窗口背景颜色

        # 设置图标
        master.iconbitmap("icon.ico")

        # 创建姓名标签
        self.name_label = tk.Label(master, font=("楷体", 120), fg="#005f35", bg=bg_color)
        self.name_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # 创建显示文件路径的标签
        self.file_path_label = tk.Label(master, text="", font=("等线", 10), fg='gray', bg=bg_color)
        self.file_path_label.place(relx=0, rely=0, anchor=tk.NW)

        # 创建导入名单按钮，改名为"导入名单"
        self.import_button = tk.Button(master, text="导入名单", command=self.manual_import)
        self.import_button.place(relx=1, rely=0, anchor=tk.NE)

        # 创建显示姓名总数的标签
        self.total_names_label = tk.Label(master, text="姓名总数：0", font=("仿宋", 10), bg=bg_color)
        self.total_names_label.place(relx=0, rely=1, anchor=tk.SW)

        # 创建音效开关按钮
        self.sound_button = tk.Button(master, text="关闭音效", command=self.toggle_sound)
        self.sound_button.place(relx=0.09, rely=0.93, anchor=tk.SE)
        self.sound_enabled = True  # 初始音效状态为开启

        # 加载姓名数据
        self.load_names_from_excel()

        # 绑定事件
        self.name_label.bind("<Button-1>", self.update_name)
        self.master.bind("<space>", self.update_name)
        self.master.bind("<Return>", self.update_name)
        self.master.bind("<Down>", self.update_name)

        # 添加全局键盘监听事件
        self.shift_pressed = False
        keyboard.on_press_key("v", self.toggle_window)

        # 添加提示文本
        self.tip_label = tk.Label(master, text="按下 V 键隐藏/显示窗口", font=("仿宋", 10), fg="gray", bg=bg_color)
        self.tip_label.place(relx=1.0, rely=1.0, anchor=tk.SE)

        
        self.version_label = tk.Label(master, text=f"当前是离线版本", font=("仿宋", 10), bg=bg_color)
        self.version_label.place(relx=0.5, rely=1, anchor=tk.SE)
        
        self.update_button = tk.Button(master, text="下载在线版", command=self.view_update, bg=bg_color)
        self.update_button.place(relx=1, rely=0.93, anchor=tk.SE)

        # 初始化pygame
        pygame.mixer.init()

        # 设置窗口置顶
        master.attributes("-topmost", True)

    # 加载姓名数据
    def load_names_from_excel(self):
        try:
            # 尝试读取上次导入的文件路径
            with open("last_file_path.txt", "r") as f:
                file_path = f.read().strip()
        except FileNotFoundError:
            file_path = None

        if file_path:
            try:
                self.names = []
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    self.names.extend(row)
                self.file_path_label.config(text=f"当前文件路径：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"加载姓名数据失败：{e}")
        else:
            self.names = []
            self.file_path_label.config(text="当前文件路径：无")

        # 更新姓名总数标签
        self.total_names_label.config(text=f"姓名总数：{len(self.names)}")

    # 选择随机姓名
    def update_name(self, event=None):
        try:
            name = random.choice(self.names)
            self.name_label.config(text=name)
            # 播放音效
            if self.sound_enabled:
                pygame.mixer.music.load("sound.wav")
                pygame.mixer.music.play()
        except Exception as e:
            messagebox.showerror("错误", f"出现错误：{e}")

    # 切换隐藏/显示窗口状态
    def toggle_window(self, event):
        self.shift_pressed = not self.shift_pressed
        if self.shift_pressed:
            self.master.withdraw()  # 隐藏窗口
        else:
            self.master.deiconify()  # 显示窗口

    # 手动导入功能
    def manual_import(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
            if file_path:
                self.names = []
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    self.names.extend(row)
                self.save_last_file_path(file_path)
                self.update_name()  # 导入成功后刷新姓名
                self.file_path_label.config(text=f"当前文件路径：{file_path}")
                self.total_names_label.config(text=f"姓名总数：{len(self.names)}")  # 更新姓名总数
                messagebox.showinfo("提示", "导入成功！")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{e}")

    # 保存上次导入的文件路径
    def save_last_file_path(self, file_path):
        try:
            with open("last_file_path.txt", "w") as f:
                f.write(file_path)
        except Exception as e:
            messagebox.showerror("错误", f"保存文件路径失败：{e}")

    # 打开最新版本的 release 页面
    def view_update(self):
        webbrowser.open("https://github.com/HowCam/ClickToChooseARandomName/releases/latest")

    # 切换音效状态
    def toggle_sound(self):
        self.sound_enabled = not self.sound_enabled
        if self.sound_enabled:
            self.sound_button.config(text="关闭音效")
        else:
            self.sound_button.config(text="开启音效")

def main():
    root = tk.Tk()
    app = FloatingWindow(root)
    app.update_name()  # 初始化时更新一次姓名
    root.mainloop()

if __name__ == "__main__":
    main()